"""
Excel workbook writer for backtest results.

Six sheets:
  1. Trade Log       — every closed trade across all phases
  2. Summary         — per-phase statistics table
  3. Monthly Returns — heat-map grid of monthly PnL % per phase
  4. Confluence Dist — trade count + avg PnL by confluence score
  5. Config Used     — all settings snapshot used in this run
  6. Equity Curve    — running balance per phase over time
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backtest.simulator import Portfolio

# ── Colour palette (dark theme) ────────────────────────────────────────────────
_HEADER_BG = "16213e"
_ROW_BG    = "111827"
_WIN_BG    = "14532d"
_LOSS_BG   = "7f1d1d"
_NEUT_BG   = "1e1b4b"
_WHITE     = "ffffff"

# One accent colour per phase: indigo / cyan / amber
_PHASE_BG = ["4338ca", "0e7490", "b45309"]

_HEADER_FONT  = Font(bold=True, color=_WHITE, name="Calibri", size=10)
_NORMAL_FONT  = Font(color=_WHITE, name="Calibri", size=10)
_MONO_FONT    = Font(color=_WHITE, name="Consolas", size=10)


def _fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def _thin_border() -> Border:
    s = Side(style="thin", color="2d2d4e")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_row(
    ws,
    row: int,
    values: list,
    *,
    bold: bool = False,
    bg: str | None = None,
    mono: bool = False,
) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")
        if bold:
            cell.font = _HEADER_FONT
        elif mono:
            cell.font = _MONO_FONT
        else:
            cell.font = _NORMAL_FONT
        if bg:
            cell.fill = _fill(bg)


def _auto_width(ws, max_width: int = 32) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


# ── Sheet 1: Trade Log ─────────────────────────────────────────────────────────

def _trade_log(ws, portfolios: list[Portfolio]) -> None:
    ws.title = "Trade Log"
    headers = [
        "Phase", "Entry Time", "Exit Time", "Action",
        "Entry Price", "Exit Price", "Stop Price", "TP Price",
        "Position USD", "PnL USD", "PnL %", "Exit Reason",
        "Balance After", "Confluence", "Rule", "LLM Conf",
    ]
    _write_row(ws, 1, headers, bold=True, bg=_HEADER_BG)

    row = 2
    for p in portfolios:
        for t in sorted(p.closed, key=lambda x: x.entry_time):
            bg = _WIN_BG if t.pnl_usd > 0 else _LOSS_BG if t.pnl_usd < 0 else _NEUT_BG
            _write_row(ws, row, [
                t.phase,
                t.entry_time.strftime("%Y-%m-%d %H:%M"),
                t.exit_time.strftime("%Y-%m-%d %H:%M"),
                t.action.upper(),
                round(t.entry_price, 2),
                round(t.exit_price, 2),
                round(t.stop_price, 2),
                round(t.take_profit_price, 2),
                round(t.position_usd, 2),
                t.pnl_usd,
                f"{t.pnl_pct:+.3f}%",
                t.exit_reason,
                t.balance_after,
                t.confluence_score if t.confluence_score is not None else "-",
                t.rule_triggered or "-",
                round(t.llm_confidence, 3) if t.llm_confidence is not None else "-",
            ], bg=bg, mono=True)
            row += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


# ── Sheet 2: Summary ───────────────────────────────────────────────────────────

_SUMMARY_METRICS: list[tuple[str, str]] = [
    ("Total Trades",           "total_trades"),
    ("Wins",                   "wins"),
    ("Losses",                 "losses"),
    ("Win Rate %",             "win_rate_pct"),
    ("Avg Win %",              "avg_win_pct"),
    ("Avg Loss %",             "avg_loss_pct"),
    ("Profit Factor",          "profit_factor"),
    ("Expectancy %",           "expectancy_pct"),
    ("Total Return %",         "total_return_pct"),
    ("Final Balance USD",      "final_balance_usd"),
    ("Max Drawdown %",         "max_drawdown_pct"),
    ("Sharpe (annualised)",    "sharpe_annualised"),
    ("Break-Even Win Rate %",  "break_even_wr_pct"),
    ("Stop Loss Exits",        "stop_hits"),
    ("Take Profit Exits",      "tp_hits"),
    ("End-of-Data Exits",      "eod_hits"),
    ("Avg Trade Duration (h)", "avg_duration_hours"),
]


def _summary(ws, portfolios: list[Portfolio]) -> None:
    ws.title = "Summary"
    headers = ["Metric"] + [f"Phase {p.phase}" for p in portfolios]
    _write_row(ws, 1, headers, bold=True, bg=_HEADER_BG)

    stats_list = [p.compute_stats() for p in portfolios]

    for r, (label, key) in enumerate(_SUMMARY_METRICS, 2):
        row_vals: list = [label]
        for i, s in enumerate(stats_list):
            v = s.get(key, "-")
            row_vals.append(v)

        # Colour total return row
        highlight: str | None = None
        if key == "total_return_pct":
            highlight = _WIN_BG if any(
                isinstance(s.get(key), (int, float)) and s.get(key, 0) > 0
                for s in stats_list
            ) else _LOSS_BG

        _write_row(ws, r, row_vals, bg=highlight or _ROW_BG)

    _auto_width(ws)
    ws.freeze_panes = "B2"


# ── Sheet 3: Monthly Returns ───────────────────────────────────────────────────

_MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]


def _monthly_returns(ws, portfolios: list[Portfolio]) -> None:
    ws.title = "Monthly Returns"

    all_monthly = [p.monthly_returns() for p in portfolios]
    all_keys: set[tuple[int, int]] = set()
    for m in all_monthly:
        all_keys.update(m.keys())

    if not all_keys:
        ws.cell(1, 1, "No closed trades").font = _NORMAL_FONT
        return

    sorted_keys = sorted(all_keys)

    # Header
    ws.cell(1, 1, "Year / Month").font = _HEADER_FONT
    ws.cell(1, 1).fill = _fill(_HEADER_BG)
    ws.cell(1, 1).border = _thin_border()
    for col_i, p in enumerate(portfolios, 2):
        cell = ws.cell(1, col_i, f"Phase {p.phase}")
        cell.font = _HEADER_FONT
        cell.fill = _fill(_PHASE_BG[p.phase - 1])
        cell.border = _thin_border()

    # Data rows
    for row_i, (yr, mo) in enumerate(sorted_keys, 2):
        label_cell = ws.cell(row_i, 1, f"{yr}-{_MONTH_ABBR[mo - 1]}")
        label_cell.font = _NORMAL_FONT
        label_cell.fill = _fill(_ROW_BG)
        label_cell.border = _thin_border()

        for col_i, monthly in enumerate(all_monthly, 2):
            val = monthly.get((yr, mo), 0.0)
            cell = ws.cell(row_i, col_i, round(val, 3))
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if val > 0:
                cell.fill = _fill(_WIN_BG)
                cell.font = Font(color="4ade80", name="Calibri", size=10)
            elif val < 0:
                cell.fill = _fill(_LOSS_BG)
                cell.font = Font(color="f87171", name="Calibri", size=10)
            else:
                cell.fill = _fill(_ROW_BG)
                cell.font = _NORMAL_FONT

    _auto_width(ws)
    ws.freeze_panes = "B2"


# ── Sheet 4: Confluence Distribution ──────────────────────────────────────────

def _confluence_dist(ws, portfolios: list[Portfolio]) -> None:
    ws.title = "Confluence Dist"

    trade_hdrs = [f"Ph{p.phase} Trades" for p in portfolios]
    pnl_hdrs   = [f"Ph{p.phase} Avg PnL%" for p in portfolios]
    headers    = ["Score"] + trade_hdrs + pnl_hdrs
    _write_row(ws, 1, headers, bold=True, bg=_HEADER_BG)

    for row_i, sc in enumerate(range(9), 2):
        row_vals: list = [sc]
        for p in portfolios:
            trades = [t for t in p.closed if t.confluence_score == sc]
            row_vals.append(len(trades))
        for p in portfolios:
            trades = [t for t in p.closed if t.confluence_score == sc]
            avg_pnl = (
                sum(t.pnl_pct for t in trades) / len(trades)
                if trades else 0.0
            )
            row_vals.append(round(avg_pnl, 3))
        _write_row(ws, row_i, row_vals, bg=_ROW_BG)

    _auto_width(ws)


# ── Sheet 5: Config Used ───────────────────────────────────────────────────────

def _config_sheet(ws, config: dict[str, Any]) -> None:
    ws.title = "Config Used"
    _write_row(ws, 1, ["Parameter", "Value"], bold=True, bg=_HEADER_BG)
    for r, (k, v) in enumerate(config.items(), 2):
        _write_row(ws, r, [k, str(v)], bg=_ROW_BG)
    _auto_width(ws)


# ── Sheet 6: Equity Curve ──────────────────────────────────────────────────────

def _equity_curve(ws, portfolios: list[Portfolio]) -> None:
    ws.title = "Equity Curve"
    headers = ["Timestamp"] + [f"Phase {p.phase} Balance" for p in portfolios]
    _write_row(ws, 1, headers, bold=True, bg=_HEADER_BG)

    # Collect all timestamps and forward-fill balances per phase
    last_bal = {p.phase: p.starting_balance for p in portfolios}
    phase_events: dict[datetime, dict[int, float]] = {}
    for p in portfolios:
        for ts, bal in p.equity_curve:
            phase_events.setdefault(ts, {})[p.phase] = bal

    for row_i, ts in enumerate(sorted(phase_events.keys()), 2):
        for p in portfolios:
            if p.phase in phase_events[ts]:
                last_bal[p.phase] = phase_events[ts][p.phase]
        row_vals: list = [ts.strftime("%Y-%m-%d %H:%M")]
        row_vals += [last_bal[p.phase] for p in portfolios]
        _write_row(ws, row_i, row_vals, bg=_ROW_BG, mono=True)

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ── Public entry point ─────────────────────────────────────────────────────────

def write_excel(
    portfolios: list[Portfolio],
    config: dict[str, Any],
    output_path: Path,
) -> Path:
    """
    Write a complete backtest report to an Excel workbook.

    Args:
        portfolios:  List of Portfolio objects (one per phase), in phase order.
        config:      Settings snapshot used in this run.
        output_path: Destination .xlsx file path.

    Returns:
        Resolved absolute output_path.
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove the default blank sheet

    ws_log     = wb.create_sheet()
    ws_summary = wb.create_sheet()
    ws_monthly = wb.create_sheet()
    ws_conf    = wb.create_sheet()
    ws_cfg     = wb.create_sheet()
    ws_equity  = wb.create_sheet()

    _trade_log(ws_log, portfolios)
    _summary(ws_summary, portfolios)
    _monthly_returns(ws_monthly, portfolios)
    _confluence_dist(ws_conf, portfolios)
    _config_sheet(ws_cfg, config)
    _equity_curve(ws_equity, portfolios)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path.resolve()
